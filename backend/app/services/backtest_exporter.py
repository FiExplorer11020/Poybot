from __future__ import annotations

import json
import re
from collections.abc import Mapping
from datetime import UTC, datetime
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from typing import Any
from uuid import uuid4

from openpyxl import Workbook
from openpyxl.chart import AreaChart, LineChart, Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

PROJECT_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_RESULTS_DIR = PROJECT_ROOT / "data" / "backtests"

BACKGROUND_HEX = "1A1A2E"
SURFACE_HEX = "16213E"
HEADER_HEX = "0F3460"
VIOLET_HEX = "7C3AED"
CYAN_HEX = "22D3EE"
SUCCESS_HEX = "22C55E"
DRAWDOWN_HEX = "EF4444"
TEXT_HEX = "F8FAFC"
MUTED_HEX = "94A3B8"
BORDER_HEX = "334155"

CURRENCY_FORMAT = "$#,##0.00;[Red]($#,##0.00);-"
PERCENT_FORMAT = "0.00%;[Red](0.00%);-"
SIZE_FORMAT = "#,##0.0000;[Red](#,##0.0000);-"

KNOWN_RESULT_KEYS = {
    "backtest_id",
    "total_trades",
    "winning_trades",
    "win_rate",
    "total_pnl",
    "total_pnl_pct",
    "max_drawdown",
    "max_drawdown_pct",
    "sharpe_ratio",
    "profit_factor",
    "avg_trade_duration_h",
    "equity_curve",
    "trades",
    "parameters",
    "params",
    "config",
    "settings",
    "metadata",
}

THIN_SIDE = Side(style="thin", color=BORDER_HEX)
DEFAULT_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)

BODY_FONT = Font(name="Calibri", size=11, color=TEXT_HEX)
BODY_BOLD_FONT = Font(name="Calibri", size=11, bold=True, color=TEXT_HEX)
SUBTITLE_FONT = Font(name="Calibri", size=10, italic=True, color=MUTED_HEX)
LOGO_FONT = Font(name="Calibri", size=18, bold=True, color=TEXT_HEX)
TITLE_FONT = Font(name="Calibri", size=22, bold=True, color=TEXT_HEX)
SECTION_FONT = Font(name="Calibri", size=13, bold=True, color=TEXT_HEX)
KPI_LABEL_FONT = Font(name="Calibri", size=10, bold=True, color=TEXT_HEX)
KPI_VALUE_FONT = Font(name="Calibri", size=18, bold=True, color=TEXT_HEX)
POSITIVE_FONT = Font(name="Calibri", size=11, bold=True, color=SUCCESS_HEX)
NEGATIVE_FONT = Font(name="Calibri", size=11, bold=True, color=DRAWDOWN_HEX)

BACKGROUND_FILL = PatternFill(fill_type="solid", fgColor=BACKGROUND_HEX)
SURFACE_FILL = PatternFill(fill_type="solid", fgColor=SURFACE_HEX)
HEADER_FILL = PatternFill(fill_type="solid", fgColor=HEADER_HEX)
VIOLET_FILL = PatternFill(fill_type="solid", fgColor=VIOLET_HEX)
CYAN_FILL = PatternFill(fill_type="solid", fgColor=CYAN_HEX)
TEAL_FILL = PatternFill(fill_type="solid", fgColor="0E7490")
RED_FILL = PatternFill(fill_type="solid", fgColor="7F1D1D")
POSITIVE_FILL = PatternFill(fill_type="solid", fgColor="123524")
NEGATIVE_FILL = PatternFill(fill_type="solid", fgColor="4C1D24")


class BacktestExportError(Exception):
    """Base error for backtest export operations."""


class BacktestResultNotFoundError(BacktestExportError):
    """Raised when a stored backtest payload cannot be found."""


class BacktestResultInvalidError(BacktestExportError):
    """Raised when a stored backtest payload is malformed."""


class BacktestExportService:
    """Persist and export backtest payloads as styled Excel workbooks."""

    def __init__(self, results_dir: str | Path | None = None) -> None:
        self.results_dir = self._resolve_results_dir(results_dir)

    def save_backtest_result(
        self, result: Mapping[str, Any], backtest_id: str | None = None
    ) -> str:
        payload = dict(result)
        generated_id = self._validate_backtest_id(backtest_id or self._new_backtest_id())
        payload["backtest_id"] = generated_id

        self.results_dir.mkdir(parents=True, exist_ok=True)
        destination = self.results_dir / f"{generated_id}.json"
        destination.write_text(
            json.dumps(payload, ensure_ascii=True, indent=2, default=self._json_default),
            encoding="utf-8",
        )
        return generated_id

    def load_backtest_result(self, backtest_id: str) -> dict[str, Any]:
        safe_backtest_id = self._validate_backtest_id(backtest_id)
        result_path = self.results_dir / f"{safe_backtest_id}.json"
        if not result_path.is_file():
            raise BacktestResultNotFoundError(
                f"backtest result '{safe_backtest_id}' not found in {self.results_dir.as_posix()}"
            )

        try:
            payload = json.loads(result_path.read_text(encoding="utf-8"))
        except json.JSONDecodeError as exc:
            raise BacktestResultInvalidError(
                f"backtest result '{safe_backtest_id}' is not valid JSON"
            ) from exc

        if not isinstance(payload, Mapping):
            raise BacktestResultInvalidError(
                f"backtest result '{safe_backtest_id}' must be a JSON object"
            )

        return dict(payload)

    def export_backtest(self, backtest_id: str) -> tuple[str, bytes]:
        safe_backtest_id = self._validate_backtest_id(backtest_id)
        payload = self.load_backtest_result(safe_backtest_id)
        workbook = self.build_workbook(safe_backtest_id, payload)

        file_name = f"poybot-backtest-{self._sanitize_filename(safe_backtest_id)}.xlsx"
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return file_name, buffer.getvalue()

    def build_workbook(self, backtest_id: str, result: Mapping[str, Any]) -> Workbook:
        equity_points = self._normalize_equity_curve(result.get("equity_curve"))
        trades = self._normalize_trades(result.get("trades"))
        summary_metrics = self._extract_summary_metrics(backtest_id, result, equity_points, trades)
        parameter_rows = self._extract_parameter_rows(backtest_id, result)

        workbook = Workbook()
        workbook.properties.creator = "Poybot"
        workbook.properties.title = f"Poybot Backtest Report - {backtest_id}"
        workbook.properties.subject = "Polymarket backtest export"
        workbook.properties.description = "Professional Excel export for a Polymarket backtest"

        if hasattr(workbook, "calculation"):
            workbook.calculation.calcMode = "auto"
            workbook.calculation.fullCalcOnLoad = True
            workbook.calculation.forceFullCalc = True

        summary_sheet = workbook.active
        summary_sheet.title = "Summary"
        self._build_summary_sheet(summary_sheet, backtest_id, summary_metrics, parameter_rows)

        equity_sheet = workbook.create_sheet("Equity Curve")
        self._build_equity_sheet(equity_sheet, backtest_id, equity_points)

        trades_sheet = workbook.create_sheet("Trades")
        self._build_trades_sheet(trades_sheet, backtest_id, trades)

        return workbook

    def _build_summary_sheet(
        self,
        ws,
        backtest_id: str,
        summary_metrics: Mapping[str, Any],
        parameter_rows: list[tuple[str, str]],
    ) -> None:
        self._apply_sheet_defaults(ws, VIOLET_HEX)
        self._set_column_widths(
            ws,
            {
                "A": 16,
                "B": 16,
                "C": 16,
                "D": 16,
                "E": 18,
                "F": 18,
                "G": 18,
                "H": 18,
            },
        )
        self._fill_region(ws, 1, 32, 1, 8, BACKGROUND_FILL)

        ws.merge_cells("A1:B2")
        ws["A1"] = "POYBOT"
        self._apply_cell_style(
            ws["A1"],
            fill=VIOLET_FILL,
            font=LOGO_FONT,
            alignment=ALIGN_CENTER,
            border=DEFAULT_BORDER,
        )

        ws.merge_cells("C1:H2")
        ws["C1"] = "Backtest Report"
        self._apply_cell_style(
            ws["C1"],
            fill=HEADER_FILL,
            font=TITLE_FONT,
            alignment=ALIGN_CENTER,
            border=DEFAULT_BORDER,
        )

        ws.merge_cells("A3:H3")
        ws["A3"] = (
            f"Backtest ID: {backtest_id}  |  Generated at: "
            f"{datetime.now(UTC).strftime('%Y-%m-%d %H:%M:%S UTC')}"
        )
        self._apply_cell_style(
            ws["A3"],
            fill=SURFACE_FILL,
            font=SUBTITLE_FONT,
            alignment=ALIGN_LEFT,
            border=DEFAULT_BORDER,
        )

        self._write_kpi_box(
            ws,
            "Win Rate",
            self._display_percent(summary_metrics.get("win_rate")),
            "A6:B8",
            CYAN_FILL,
        )
        self._write_kpi_box(
            ws,
            "Total PnL",
            self._display_currency(
                summary_metrics.get("total_pnl"), percent=summary_metrics.get("total_pnl_pct")
            ),
            "C6:D8",
            VIOLET_FILL,
        )
        self._write_kpi_box(
            ws,
            "Sharpe",
            self._display_ratio(summary_metrics.get("sharpe_ratio")),
            "E6:F8",
            TEAL_FILL,
        )
        self._write_kpi_box(
            ws,
            "Max DD",
            self._display_currency(
                summary_metrics.get("max_drawdown"), percent=summary_metrics.get("max_drawdown_pct")
            ),
            "G6:H8",
            RED_FILL,
        )

        self._write_section_title(ws, "A11:B11", "Overview")
        self._write_section_title(ws, "D11:H11", "Parameters")

        overview_rows = [
            ("Total trades", self._display_integer(summary_metrics.get("total_trades"))),
            ("Winning trades", self._display_integer(summary_metrics.get("winning_trades"))),
            ("Profit factor", self._display_ratio(summary_metrics.get("profit_factor"))),
            ("Avg duration (h)", self._display_ratio(summary_metrics.get("avg_trade_duration_h"))),
            ("Total return", self._display_percent(summary_metrics.get("total_pnl_pct"))),
            ("Max drawdown %", self._display_percent(summary_metrics.get("max_drawdown_pct"))),
        ]
        self._write_key_value_table(
            ws, start_row=12, start_col=1, rows=overview_rows, header=("Metric", "Value")
        )
        self._write_key_value_table(
            ws, start_row=12, start_col=4, rows=parameter_rows, header=("Parameter", "Value")
        )

        ws.freeze_panes = "A6"

    def _build_equity_sheet(
        self, ws, backtest_id: str, equity_points: list[dict[str, Any]]
    ) -> None:
        self._apply_sheet_defaults(ws, CYAN_HEX)
        self._set_column_widths(ws, {"A": 24, "B": 16, "C": 16})
        self._fill_region(ws, 1, max(18, len(equity_points) + 8), 1, 9, BACKGROUND_FILL)

        ws.merge_cells("A1:C2")
        ws["A1"] = f"Equity Curve | {backtest_id}"
        self._apply_cell_style(
            ws["A1"],
            fill=HEADER_FILL,
            font=TITLE_FONT,
            alignment=ALIGN_CENTER,
            border=DEFAULT_BORDER,
        )

        header_row = 4
        self._write_table_header(ws, header_row, ["Timestamp", "Equity", "Drawdown %"])
        data_start_row = header_row + 1

        if not equity_points:
            ws.merge_cells(
                start_row=data_start_row, start_column=1, end_row=data_start_row, end_column=3
            )
            ws.cell(row=data_start_row, column=1, value="No equity curve points found in payload.")
            self._apply_cell_style(
                ws["A5"],
                fill=SURFACE_FILL,
                font=BODY_FONT,
                alignment=ALIGN_LEFT,
                border=DEFAULT_BORDER,
            )
            return

        for row_index, point in enumerate(equity_points, start=data_start_row):
            timestamp_cell = ws.cell(row=row_index, column=1, value=point["timestamp_value"])
            equity_cell = ws.cell(row=row_index, column=2, value=point["equity"])
            drawdown_cell = ws.cell(row=row_index, column=3, value=point["drawdown_pct"])

            self._apply_cell_style(
                timestamp_cell,
                fill=SURFACE_FILL,
                font=BODY_FONT,
                alignment=ALIGN_LEFT,
                border=DEFAULT_BORDER,
                number_format=(
                    "yyyy-mm-dd hh:mm:ss"
                    if isinstance(point["timestamp_value"], datetime)
                    else "General"
                ),
            )
            self._apply_cell_style(
                equity_cell,
                fill=SURFACE_FILL,
                font=BODY_FONT,
                alignment=ALIGN_RIGHT,
                border=DEFAULT_BORDER,
                number_format=CURRENCY_FORMAT,
            )
            self._apply_cell_style(
                drawdown_cell,
                fill=SURFACE_FILL,
                font=BODY_FONT,
                alignment=ALIGN_RIGHT,
                border=DEFAULT_BORDER,
                number_format=PERCENT_FORMAT,
            )

        data_end_row = data_start_row + len(equity_points) - 1
        ws.auto_filter.ref = f"A{header_row}:C{data_end_row}"
        ws.freeze_panes = "A5"
        self._add_equity_chart(
            ws, header_row=header_row, data_start_row=data_start_row, data_end_row=data_end_row
        )

    def _build_trades_sheet(self, ws, backtest_id: str, trades: list[dict[str, Any]]) -> None:
        self._apply_sheet_defaults(ws, VIOLET_HEX)
        self._set_column_widths(
            ws,
            {
                "A": 22,
                "B": 44,
                "C": 14,
                "D": 14,
                "E": 12,
                "F": 14,
                "G": 14,
                "H": 12,
                "I": 14,
            },
        )
        self._fill_region(ws, 1, max(20, len(trades) + 10), 1, 9, BACKGROUND_FILL)

        ws.merge_cells("A1:I2")
        ws["A1"] = f"Trades Ledger | {backtest_id}"
        self._apply_cell_style(
            ws["A1"],
            fill=HEADER_FILL,
            font=TITLE_FONT,
            alignment=ALIGN_CENTER,
            border=DEFAULT_BORDER,
        )

        header_row = 4
        self._write_table_header(
            ws,
            header_row,
            [
                "Date",
                "Marche",
                "Side",
                "Prix d'entree",
                "Taille",
                "Notional",
                "PnL $",
                "PnL %",
                "Status",
            ],
        )
        data_start_row = header_row + 1

        for row_index, trade in enumerate(trades, start=data_start_row):
            values = [
                trade["date_value"],
                trade["market_title"],
                trade["side"],
                trade["entry_price"],
                trade["size"],
                trade["notional"],
                trade["pnl_abs"],
                trade["pnl_pct"],
                trade["status"],
            ]
            for column_index, value in enumerate(values, start=1):
                cell = ws.cell(row=row_index, column=column_index, value=value)
                number_format = "General"
                alignment = ALIGN_LEFT
                if column_index == 1 and isinstance(value, datetime):
                    number_format = "yyyy-mm-dd hh:mm:ss"
                elif column_index in {4, 6, 7}:
                    number_format = CURRENCY_FORMAT
                    alignment = ALIGN_RIGHT
                elif column_index == 5:
                    number_format = SIZE_FORMAT
                    alignment = ALIGN_RIGHT
                elif column_index == 8:
                    number_format = PERCENT_FORMAT
                    alignment = ALIGN_RIGHT
                elif column_index == 9:
                    alignment = ALIGN_CENTER

                self._apply_cell_style(
                    cell,
                    fill=SURFACE_FILL,
                    font=BODY_FONT,
                    alignment=alignment,
                    border=DEFAULT_BORDER,
                    number_format=number_format,
                )

        if trades:
            data_end_row = data_start_row + len(trades) - 1
            ws.auto_filter.ref = f"A{header_row}:I{data_end_row}"
            self._apply_trade_conditional_formatting(ws, data_start_row, data_end_row)
        else:
            data_end_row = data_start_row
            ws.merge_cells(
                start_row=data_start_row, start_column=1, end_row=data_start_row, end_column=9
            )
            ws.cell(row=data_start_row, column=1, value="No trades found in payload.")
            self._apply_cell_style(
                ws["A5"],
                fill=SURFACE_FILL,
                font=BODY_FONT,
                alignment=ALIGN_LEFT,
                border=DEFAULT_BORDER,
            )

        totals_row = data_end_row + 2
        ws.merge_cells(start_row=totals_row, start_column=1, end_row=totals_row, end_column=6)
        totals_label = ws.cell(row=totals_row, column=1, value="Totals")
        self._apply_cell_style(
            totals_label,
            fill=HEADER_FILL,
            font=SECTION_FONT,
            alignment=ALIGN_LEFT,
            border=DEFAULT_BORDER,
        )

        pnl_total_cell = ws.cell(
            row=totals_row,
            column=7,
            value=f"=SUM(G{data_start_row}:G{data_end_row})" if trades else 0,
        )
        pnl_pct_avg_cell = ws.cell(
            row=totals_row,
            column=8,
            value=f"=AVERAGE(H{data_start_row}:H{data_end_row})" if trades else 0,
        )
        trade_count_cell = ws.cell(row=totals_row, column=9, value=len(trades))

        self._apply_cell_style(
            pnl_total_cell,
            fill=HEADER_FILL,
            font=BODY_BOLD_FONT,
            alignment=ALIGN_RIGHT,
            border=DEFAULT_BORDER,
            number_format=CURRENCY_FORMAT,
        )
        self._apply_cell_style(
            pnl_pct_avg_cell,
            fill=HEADER_FILL,
            font=BODY_BOLD_FONT,
            alignment=ALIGN_RIGHT,
            border=DEFAULT_BORDER,
            number_format=PERCENT_FORMAT,
        )
        self._apply_cell_style(
            trade_count_cell,
            fill=HEADER_FILL,
            font=BODY_BOLD_FONT,
            alignment=ALIGN_CENTER,
            border=DEFAULT_BORDER,
            number_format="0",
        )

        ws.freeze_panes = "A5"

    def _apply_trade_conditional_formatting(self, ws, start_row: int, end_row: int) -> None:
        for column_letter in ("G", "H"):
            data_range = f"{column_letter}{start_row}:{column_letter}{end_row}"
            ws.conditional_formatting.add(
                data_range,
                CellIsRule(
                    operator="greaterThan", formula=["0"], fill=POSITIVE_FILL, font=POSITIVE_FONT
                ),
            )
            ws.conditional_formatting.add(
                data_range,
                CellIsRule(
                    operator="lessThan", formula=["0"], fill=NEGATIVE_FILL, font=NEGATIVE_FONT
                ),
            )

    def _add_equity_chart(
        self, ws, header_row: int, data_start_row: int, data_end_row: int
    ) -> None:
        categories = Reference(ws, min_col=1, min_row=data_start_row, max_row=data_end_row)
        equity_data = Reference(ws, min_col=2, min_row=header_row, max_row=data_end_row)
        drawdown_data = Reference(ws, min_col=3, min_row=header_row, max_row=data_end_row)

        line_chart = LineChart()
        line_chart.title = "Equity Curve"
        line_chart.style = 2
        line_chart.height = 9
        line_chart.width = 22
        line_chart.y_axis.title = "Equity ($)"
        line_chart.x_axis.title = "Timestamp"
        line_chart.x_axis.tickLblPos = "low"
        line_chart.legend.position = "r"
        line_chart.add_data(equity_data, titles_from_data=True)
        line_chart.set_categories(categories)

        if line_chart.series:
            equity_series = line_chart.series[0]
            equity_series.graphicalProperties.line.solidFill = CYAN_HEX
            equity_series.graphicalProperties.line.width = 28575

        drawdown_chart = AreaChart()
        drawdown_chart.style = 13
        drawdown_chart.grouping = "standard"
        drawdown_chart.add_data(drawdown_data, titles_from_data=True)
        drawdown_chart.set_categories(categories)
        drawdown_chart.y_axis.title = "Drawdown %"
        drawdown_chart.y_axis.axId = 200
        drawdown_chart.y_axis.crosses = "max"
        drawdown_chart.y_axis.majorGridlines = None

        if drawdown_chart.series:
            drawdown_series = drawdown_chart.series[0]
            drawdown_series.graphicalProperties.solidFill = DRAWDOWN_HEX
            drawdown_series.graphicalProperties.line.solidFill = DRAWDOWN_HEX

        line_chart += drawdown_chart
        ws.add_chart(line_chart, "E4")

    def _extract_summary_metrics(
        self,
        backtest_id: str,
        result: Mapping[str, Any],
        equity_points: list[dict[str, Any]],
        trades: list[dict[str, Any]],
    ) -> dict[str, Any]:
        total_trades = self._coerce_int(result.get("total_trades"))
        winning_trades = self._coerce_int(result.get("winning_trades"))
        win_rate = self._normalize_percent(result.get("win_rate"))
        total_pnl = self._coerce_number(result.get("total_pnl"))
        total_pnl_pct = self._normalize_percent(result.get("total_pnl_pct"))
        max_drawdown = self._coerce_number(result.get("max_drawdown"))
        max_drawdown_pct = self._normalize_percent(result.get("max_drawdown_pct"))

        if total_trades is None:
            total_trades = len(trades)
        if winning_trades is None and trades:
            winning_trades = sum(1 for trade in trades if (trade.get("pnl_abs") or 0) > 0)
        if win_rate is None and trades:
            win_rate = (winning_trades or 0) / len(trades)
        if total_pnl is None:
            total_pnl = sum(trade.get("pnl_abs") or 0 for trade in trades)
        if total_pnl_pct is None and len(equity_points) >= 2:
            start_equity = equity_points[0]["equity"]
            end_equity = equity_points[-1]["equity"]
            if start_equity:
                total_pnl_pct = (end_equity - start_equity) / start_equity
        if max_drawdown is None and equity_points:
            max_drawdown = max(point.get("drawdown_abs", 0.0) for point in equity_points)
        if max_drawdown_pct is None and equity_points:
            max_drawdown_pct = max(point["drawdown_pct"] for point in equity_points)

        return {
            "backtest_id": backtest_id,
            "total_trades": total_trades,
            "winning_trades": winning_trades,
            "win_rate": win_rate,
            "total_pnl": total_pnl,
            "total_pnl_pct": total_pnl_pct,
            "max_drawdown": max_drawdown,
            "max_drawdown_pct": max_drawdown_pct,
            "sharpe_ratio": self._coerce_number(result.get("sharpe_ratio")),
            "profit_factor": self._coerce_number(result.get("profit_factor")),
            "avg_trade_duration_h": self._coerce_number(result.get("avg_trade_duration_h")),
        }

    def _extract_parameter_rows(
        self, backtest_id: str, result: Mapping[str, Any]
    ) -> list[tuple[str, str]]:
        parameter_sources: list[Mapping[str, Any]] = []
        for key in ("parameters", "params", "config", "settings"):
            candidate = result.get(key)
            if isinstance(candidate, Mapping):
                parameter_sources.append(candidate)

        metadata = result.get("metadata")
        if isinstance(metadata, Mapping):
            nested_parameters = metadata.get("parameters")
            if isinstance(nested_parameters, Mapping):
                parameter_sources.append(nested_parameters)

        extras = {
            key: value
            for key, value in result.items()
            if key not in KNOWN_RESULT_KEYS
            and isinstance(value, (Mapping, list, str, int, float, Decimal, bool))
        }
        if extras:
            parameter_sources.append(extras)

        rows: list[tuple[str, str]] = []
        for source in parameter_sources:
            rows.extend(self._flatten_mapping(source))

        if not rows:
            rows = [
                ("backtest_id", backtest_id),
                ("note", "No strategy parameters found in payload"),
            ]

        return rows

    def _normalize_equity_curve(self, raw_points: Any) -> list[dict[str, Any]]:
        if not isinstance(raw_points, list):
            return []

        normalized: list[dict[str, Any]] = []
        running_peak = 0.0
        for index, raw_point in enumerate(raw_points, start=1):
            if not isinstance(raw_point, Mapping):
                continue

            raw_timestamp = self._first_present(
                raw_point, ("timestamp", "date", "observed_at", "time")
            )
            parsed_timestamp = self._coerce_datetime(raw_timestamp)
            timestamp_value = self._to_excel_datetime(parsed_timestamp) or str(
                raw_timestamp or f"Point {index}"
            )
            equity = self._coerce_number(raw_point.get("equity"), default=0.0) or 0.0
            running_peak = max(running_peak, equity)
            drawdown_pct = (
                self._normalize_percent(raw_point.get("drawdown_pct"), default=0.0) or 0.0
            )
            normalized.append(
                {
                    "timestamp_value": timestamp_value,
                    "equity": equity,
                    "drawdown_pct": drawdown_pct,
                    "drawdown_abs": max(0.0, running_peak - equity),
                }
            )

        return normalized

    def _normalize_trades(self, raw_trades: Any) -> list[dict[str, Any]]:
        if not isinstance(raw_trades, list):
            return []

        normalized: list[dict[str, Any]] = []
        for index, raw_trade in enumerate(raw_trades, start=1):
            if not isinstance(raw_trade, Mapping):
                continue

            raw_date = self._first_present(
                raw_trade,
                ("date", "timestamp", "executed_at", "entry_time", "entry_timestamp", "opened_at"),
            )
            parsed_date = self._coerce_datetime(raw_date)
            date_value = self._to_excel_datetime(parsed_date) or str(raw_date or f"Trade {index}")

            entry_price = self._coerce_number(
                self._first_present(raw_trade, ("entry_price", "price", "avg_entry_price"))
            )
            size = self._coerce_number(self._first_present(raw_trade, ("size", "qty", "quantity")))
            notional = self._coerce_number(
                self._first_present(raw_trade, ("notional", "notional_usd", "value"))
            )
            if notional is None and entry_price is not None and size is not None:
                notional = entry_price * size

            normalized.append(
                {
                    "date_value": date_value,
                    "market_title": str(
                        self._first_present(
                            raw_trade,
                            ("market_title", "market", "market_name", "title", "question"),
                        )
                        or "Unknown market"
                    ),
                    "side": str(raw_trade.get("side") or raw_trade.get("direction") or "-"),
                    "entry_price": entry_price,
                    "size": size,
                    "notional": notional,
                    "pnl_abs": self._coerce_number(
                        self._first_present(
                            raw_trade, ("pnl_abs", "pnl", "realized_pnl", "net_pnl")
                        ),
                        default=0.0,
                    ),
                    "pnl_pct": self._normalize_percent(
                        self._first_present(
                            raw_trade, ("pnl_pct", "return_pct", "return_percent", "roi")
                        )
                    ),
                    "status": str(
                        raw_trade.get("status") or raw_trade.get("settlement") or "CLOSED"
                    ).upper(),
                }
            )

        return normalized

    def _flatten_mapping(self, data: Mapping[str, Any], prefix: str = "") -> list[tuple[str, str]]:
        rows: list[tuple[str, str]] = []
        for key in sorted(data):
            value = data[key]
            field_name = f"{prefix}.{key}" if prefix else str(key)
            if isinstance(value, Mapping):
                rows.extend(self._flatten_mapping(value, field_name))
            else:
                rows.append((field_name, self._stringify_value(value)))
        return rows

    def _write_kpi_box(
        self, ws, title: str, value: str, cell_range: str, fill: PatternFill
    ) -> None:
        start_cell, end_cell = cell_range.split(":")
        start_row = ws[start_cell].row
        end_row = ws[end_cell].row
        start_col = ws[start_cell].column
        end_col = ws[end_cell].column

        self._fill_region(ws, start_row, end_row, start_col, end_col, fill)
        self._border_region(ws, start_row, end_row, start_col, end_col)

        ws.merge_cells(
            start_row=start_row, start_column=start_col, end_row=start_row, end_column=end_col
        )
        label_cell = ws.cell(row=start_row, column=start_col, value=title)
        self._apply_cell_style(
            label_cell,
            fill=fill,
            font=KPI_LABEL_FONT,
            alignment=ALIGN_CENTER,
            border=DEFAULT_BORDER,
        )

        ws.merge_cells(
            start_row=start_row + 1, start_column=start_col, end_row=end_row, end_column=end_col
        )
        value_cell = ws.cell(row=start_row + 1, column=start_col, value=value)
        self._apply_cell_style(
            value_cell,
            fill=fill,
            font=KPI_VALUE_FONT,
            alignment=ALIGN_CENTER_WRAP,
            border=DEFAULT_BORDER,
        )

    def _write_section_title(self, ws, merged_range: str, title: str) -> None:
        ws.merge_cells(merged_range)
        cell = ws[merged_range.split(":")[0]]
        cell.value = title
        self._apply_cell_style(
            cell, fill=HEADER_FILL, font=SECTION_FONT, alignment=ALIGN_LEFT, border=DEFAULT_BORDER
        )

    def _write_key_value_table(
        self,
        ws,
        start_row: int,
        start_col: int,
        rows: list[tuple[str, str]],
        header: tuple[str, str],
    ) -> None:
        header_cells = [
            ws.cell(row=start_row, column=start_col + offset, value=value)
            for offset, value in enumerate(header)
        ]
        for cell in header_cells:
            self._apply_cell_style(
                cell,
                fill=VIOLET_FILL,
                font=BODY_BOLD_FONT,
                alignment=ALIGN_CENTER,
                border=DEFAULT_BORDER,
            )

        for row_offset, (label, value) in enumerate(rows, start=1):
            label_cell = ws.cell(row=start_row + row_offset, column=start_col, value=label)
            value_cell = ws.cell(row=start_row + row_offset, column=start_col + 1, value=value)
            self._apply_cell_style(
                label_cell,
                fill=SURFACE_FILL,
                font=BODY_BOLD_FONT,
                alignment=ALIGN_LEFT_WRAP,
                border=DEFAULT_BORDER,
            )
            self._apply_cell_style(
                value_cell,
                fill=SURFACE_FILL,
                font=BODY_FONT,
                alignment=ALIGN_LEFT_WRAP,
                border=DEFAULT_BORDER,
            )

    def _write_table_header(self, ws, row_index: int, headers: list[str]) -> None:
        for column_index, header in enumerate(headers, start=1):
            cell = ws.cell(row=row_index, column=column_index, value=header)
            self._apply_cell_style(
                cell,
                fill=VIOLET_FILL,
                font=BODY_BOLD_FONT,
                alignment=ALIGN_CENTER,
                border=DEFAULT_BORDER,
            )

    def _apply_sheet_defaults(self, ws, tab_color: str) -> None:
        ws.sheet_view.showGridLines = False
        ws.sheet_properties.tabColor = tab_color
        ws.sheet_format.defaultRowHeight = 20

    def _apply_cell_style(
        self,
        cell,
        *,
        fill: PatternFill,
        font: Font,
        alignment: Alignment,
        border: Border,
        number_format: str | None = None,
    ) -> None:
        cell.fill = fill
        cell.font = font
        cell.alignment = alignment
        cell.border = border
        if number_format is not None:
            cell.number_format = number_format

    def _fill_region(
        self, ws, start_row: int, end_row: int, start_col: int, end_col: int, fill: PatternFill
    ) -> None:
        for row in range(start_row, end_row + 1):
            for column in range(start_col, end_col + 1):
                ws.cell(row=row, column=column).fill = fill

    def _border_region(
        self, ws, start_row: int, end_row: int, start_col: int, end_col: int
    ) -> None:
        for row in range(start_row, end_row + 1):
            for column in range(start_col, end_col + 1):
                ws.cell(row=row, column=column).border = DEFAULT_BORDER

    def _set_column_widths(self, ws, widths: Mapping[str, float]) -> None:
        for column, width in widths.items():
            ws.column_dimensions[column].width = width

    def _first_present(self, mapping: Mapping[str, Any], keys: tuple[str, ...]) -> Any:
        for key in keys:
            value = mapping.get(key)
            if value is not None and value != "":
                return value
        return None

    def _coerce_number(self, value: Any, default: float | None = None) -> float | None:
        if value is None or value == "":
            return default
        if isinstance(value, bool):
            return float(value)
        if isinstance(value, (int, float, Decimal)):
            return float(value)
        if isinstance(value, str):
            cleaned = value.strip().replace(",", "")
            is_percent = cleaned.endswith("%")
            if is_percent:
                cleaned = cleaned[:-1]
            if cleaned.startswith("(") and cleaned.endswith(")"):
                cleaned = f"-{cleaned[1:-1]}"
            cleaned = cleaned.replace("$", "")
            try:
                parsed = float(cleaned)
            except ValueError:
                return default
            return parsed / 100 if is_percent else parsed
        return default

    def _normalize_percent(self, value: Any, default: float | None = None) -> float | None:
        number = self._coerce_number(value, default=default)
        if number is None:
            return default
        return number / 100 if abs(number) > 1 else number

    def _coerce_int(self, value: Any) -> int | None:
        number = self._coerce_number(value)
        if number is None:
            return None
        return int(number)

    def _coerce_datetime(self, value: Any) -> datetime | None:
        if value is None or value == "":
            return None
        if isinstance(value, datetime):
            return value
        if isinstance(value, (int, float)):
            return datetime.fromtimestamp(float(value), tz=UTC)
        if isinstance(value, str):
            cleaned = value.strip()
            if cleaned.endswith("Z"):
                cleaned = f"{cleaned[:-1]}+00:00"
            try:
                return datetime.fromisoformat(cleaned)
            except ValueError:
                return None
        return None

    def _to_excel_datetime(self, value: datetime | None) -> datetime | None:
        if value is None:
            return None
        if value.tzinfo is None:
            return value
        return value.astimezone(UTC).replace(tzinfo=None)

    def _display_percent(self, value: Any) -> str:
        normalized = self._normalize_percent(value)
        return "-" if normalized is None else f"{normalized * 100:,.2f}%"

    def _display_currency(self, value: Any, percent: Any | None = None) -> str:
        number = self._coerce_number(value)
        if number is None:
            return "-"
        rendered = f"${number:,.2f}"
        if percent is None:
            return rendered
        normalized_percent = self._normalize_percent(percent)
        if normalized_percent is None:
            return rendered
        return f"{rendered}\n({normalized_percent * 100:,.2f}%)"

    def _display_ratio(self, value: Any) -> str:
        number = self._coerce_number(value)
        return "-" if number is None else f"{number:,.2f}"

    def _display_integer(self, value: Any) -> str:
        number = self._coerce_int(value)
        return "-" if number is None else f"{number:,}"

    def _stringify_value(self, value: Any) -> str:
        if value is None:
            return "-"
        if isinstance(value, bool):
            return "true" if value else "false"
        if isinstance(value, (int, float, Decimal)):
            return str(value)
        if isinstance(value, str):
            return value
        if isinstance(value, list):
            return ", ".join(self._stringify_value(item) for item in value)
        return json.dumps(value, ensure_ascii=True, sort_keys=True, default=self._json_default)

    def _resolve_results_dir(self, results_dir: str | Path | None) -> Path:
        if results_dir is None:
            return DEFAULT_RESULTS_DIR

        candidate = Path(results_dir)
        if not candidate.is_absolute():
            candidate = PROJECT_ROOT / candidate
        return candidate.resolve()

    def _validate_backtest_id(self, backtest_id: str) -> str:
        if not re.fullmatch(r"[A-Za-z0-9._-]+", backtest_id):
            raise BacktestResultInvalidError(
                "backtest_id may only contain letters, numbers, '.', '_' and '-'"
            )
        return backtest_id

    def _sanitize_filename(self, value: str) -> str:
        sanitized = re.sub(r"[^A-Za-z0-9._-]+", "-", value).strip("-")
        return sanitized or "backtest"

    def _new_backtest_id(self) -> str:
        return f"bt-{datetime.now(UTC).strftime('%Y%m%d%H%M%S')}-{uuid4().hex[:8]}"

    @staticmethod
    def _json_default(value: Any) -> Any:
        if isinstance(value, datetime):
            return value.isoformat()
        if isinstance(value, Decimal):
            return float(value)
        if hasattr(value, "__dict__"):
            return value.__dict__
        raise TypeError(f"Object of type {type(value).__name__} is not JSON serializable")
