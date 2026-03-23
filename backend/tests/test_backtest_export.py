import json
from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.api.v1 import routes
from app.main import app


def test_backtest_export_returns_professional_workbook(monkeypatch, tmp_path) -> None:
    backtest_dir = tmp_path / "backtests"
    backtest_dir.mkdir()
    (backtest_dir / "bt-20260323.json").write_text(
        json.dumps(
            {
                "backtest_id": "bt-20260323",
                "total_trades": 2,
                "winning_trades": 1,
                "win_rate": 50.0,
                "total_pnl": 12.34,
                "total_pnl_pct": 1.234,
                "max_drawdown": 5.67,
                "max_drawdown_pct": 0.89,
                "sharpe_ratio": 2.41,
                "profit_factor": 3.12,
                "avg_trade_duration_h": 1.75,
                "parameters": {
                    "starting_equity": 1000,
                    "fee_bps": 10,
                    "strategy": {"min_edge": 0.02},
                },
                "equity_curve": [
                    {
                        "timestamp": "2026-03-23T09:00:00+00:00",
                        "equity": 1000.0,
                        "drawdown_pct": 0.0,
                    },
                    {
                        "timestamp": "2026-03-23T10:00:00+00:00",
                        "equity": 1012.34,
                        "drawdown_pct": 0.005,
                    },
                ],
                "trades": [
                    {
                        "entry_time": "2026-03-23T09:15:00+00:00",
                        "question": "Will BTC close above 100k this week?",
                        "side": "BUY_YES",
                        "entry_price": 0.44,
                        "size": 50,
                        "notional": 22,
                        "pnl": 4.4,
                        "pnl_pct": 20.0,
                        "settlement": "resolved",
                    },
                    {
                        "entry_time": "2026-03-23T09:45:00+00:00",
                        "question": "Will ETH ETF flows stay positive today?",
                        "side": "BUY_NO",
                        "entry_price": 0.57,
                        "size": 20,
                        "notional": 11.4,
                        "pnl": -1.25,
                        "pnl_pct": -10.96,
                        "settlement": "mtm",
                    },
                ],
            }
        ),
        encoding="utf-8",
    )

    monkeypatch.setattr(routes.settings, "backtest_results_dir", str(backtest_dir))

    with TestClient(app) as client:
        response = client.get("/api/v1/backtest/bt-20260323/export")

    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert 'filename="poybot-backtest-bt-20260323.xlsx"' in response.headers["content-disposition"]

    buffer = BytesIO(response.content)
    workbook = load_workbook(buffer)
    try:
        assert workbook.sheetnames == ["Summary", "Equity Curve", "Trades"]
        assert workbook["Summary"]["A1"].value == "POYBOT"
        assert workbook["Summary"]["C1"].value == "Backtest Report"
        assert workbook["Equity Curve"]._charts
        assert workbook["Trades"].auto_filter.ref == "A4:I6"
        assert workbook["Trades"]["A8"].value == "Totals"
    finally:
        workbook.close()
        buffer.close()


def test_backtest_export_returns_404_when_result_is_missing(monkeypatch, tmp_path) -> None:
    monkeypatch.setattr(routes.settings, "backtest_results_dir", str(tmp_path))

    with TestClient(app) as client:
        response = client.get("/api/v1/backtest/missing/export")

    assert response.status_code == 404
